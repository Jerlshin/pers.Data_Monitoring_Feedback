from django.shortcuts import render, redirect
from .forms import UserInputForm
from .models import UserInput
from django.utils import timezone
from django.db.models import Sum
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


# Login View
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('home')  # Redirect to home page after successful login
        else:
            messages.error(request, 'Invalid username or password')  # Show error message
    
    return render(request, 'monitor/login.html')  # Render the login page


# Logout View
def logout_view(request):
    logout(request)
    return redirect('login')  # Redirect to login page after logout


# Home View (after login)
@login_required
def home_view(request):
    return render(request, 'monitor/home.html')


# Input View
@login_required  # Ensure that only authenticated users can submit data
def input_view(request):
    if request.method == 'POST':
        form = UserInputForm(request.POST)
        if form.is_valid():
            user_input = form.save(commit=False)
            user_input.date = timezone.now().date()  # Set the current date
            user_input.save()
            return redirect('home')  # Redirect to the home page after saving the form
    else:
        form = UserInputForm()

    current_date = timezone.now().date()  # Get today's date
    return render(request, 'monitor/input.html', {'form': form, 'current_date': current_date})

# Chart View
@login_required  # Ensure that only logged-in users can access the chart page
def chart_view(request):
    # Retrieve all UserInput records
    inputs = UserInput.objects.all().order_by('date')  # Order by date for the time series

    # If no inputs are available, set default values to avoid errors
    if not inputs:
        line_chart_data = {'labels': [], 'cumulative_score': []}
        bar_chart_data = {'labels': [], 'values': []}
        return render(request, 'monitor/chart.html', {
            'line_chart_data': line_chart_data,
            'bar_chart_data': bar_chart_data,
        })

    # Prepare cumulative data for the line chart
    date_labels = []
    cumulative_data = []
    current_score = 0  # Start with a score of 0

    # Prepare data for the bar chart
    total_read_bible = 0
    total_prayed = 0
    total_exercised = 0
    total_satisfied = 0
    total_hurt_someone = 0

    for user_input in inputs:
        date_labels.append(user_input.date.strftime('%Y-%m-%d'))  # Use the date field for x-axis labels

        # Calculate cumulative score based on user inputs
        if user_input.read_bible:
            current_score += 1
            total_read_bible += 1
        if user_input.prayed:
            current_score += 1
            total_prayed += 1
        if user_input.exercised:
            current_score += 1
            total_exercised += 1
        if user_input.satisfied:
            current_score += 1
            total_satisfied += 1
        if user_input.hurt_someone:
            current_score -= 1  # Decrease the score if the user hurt someone
            total_hurt_someone += 1

        cumulative_data.append(current_score)  # Append the current score to the data list

    line_chart_data = {
        'labels': date_labels,
        'cumulative_score': cumulative_data,
    }

    # Prepare bar chart data by summing individual inputs
    bar_chart_data = {
        'labels': ['Read Bible', 'Prayed', 'Exercised', 'Satisfied', 'Hurt Someone'],
        'values': [
            total_read_bible,
            total_prayed,
            total_exercised,
            total_satisfied,
            -total_hurt_someone  # Show hurt as a negative value
        ],
    }

    return render(request, 'monitor/chart.html', {
        'line_chart_data': line_chart_data,
        'bar_chart_data': bar_chart_data,
    })

@login_required  # Ensure only logged-in users can export the data
def export_to_excel_view(request):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Inputs"

    # Add headers to the first row of the Excel sheet
    headers = ['Date', 'Read Bible', 'Prayed', 'Exercised', 'Satisfied', 'Hurt Someone', 'Loved Someone', 'Daily Summary']
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet[f'{column_letter}1'] = header

    # Retrieve all UserInput records and add them to the Excel sheet
    inputs = UserInput.objects.all().order_by('date')
    for row_num, user_input in enumerate(inputs, 2):
        sheet[f'A{row_num}'] = user_input.date.strftime('%Y-%m-%d')
        sheet[f'B{row_num}'] = 'Yes' if user_input.read_bible else 'No'
        sheet[f'C{row_num}'] = 'Yes' if user_input.prayed else 'No'
        sheet[f'D{row_num}'] = 'Yes' if user_input.exercised else 'No'
        sheet[f'E{row_num}'] = 'Yes' if user_input.satisfied else 'No'
        sheet[f'F{row_num}'] = 'Yes' if user_input.hurt_someone else 'No'
        sheet[f'G{row_num}'] = user_input.loved_someone  # Add Loved Someone field
        sheet[f'H{row_num}'] = user_input.daily_summary  # Add Daily Summary field

    # Create an HTTP response with the Excel file as an attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="user_inputs.xlsx"'
    workbook.save(response)
    
    return response
